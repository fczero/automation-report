#!/usr/bin/env python3

# #==============================================================================
# #      The report will be saved as an XLSX in the same directory.
# #
# #      1.      Login to the VPN
# #      2.a     Add executable permission using chmod
# #              Ex. $chmod 775 ccp_daily_automation.py
# #      3       Run using $./ccp_daily_automation.py
# #        OR
# #      2.b     Run using $python3 ccp_daily_automation.py
# #         Help
# #          usage: ccp_daily_automation.py [-h] [-s | -r | -t]
# #
# #          Scrape Jenkins report and create XLSX report.
# #
# #          optional arguments:
# #              -h, --help        show this help message and exit
# #              -s, --smoke       Generate Smoke Report
# #              -r, --regression  Generate Regression Report
# #              -t, --test        Test mode, Smoke Confirmation
# #
# #==============================================================================

import os
import sys
import subprocess
import multiprocessing as mp
import venv
import argparse
from pprint import pprint

# Global Values
# WIP
BUILD_NO         = ''
# WIP END
HEADER_ROW       = 1
LINE_WRAP_LEN    = 60
DURATION_COL_LEN = 18
MAN_TEST_COL_LEN = 18
automationReport = ''
reportFileName   = ''
suite            = {}
devices          = []
nodes            = []
ENV_NAME         = 'myenv'

#1-based indexing
COLS = ["",
        "TEST_NAME",
        "SUCCESS_RATE",
        "TEST_NO",
        "FAILED_NO",
        "DURATION",
        "FAILED_SCENARIOS",
        "FAILED_STEPS",
        "SKIPPED_SCENARIOS",
        "MANUAL_TEST"]


# janky bootsrap helper funcs
def arghandler():
    global automationReport
    global suite
    global devices
    global nodes
    global reportFileName
    fName = os.path.basename(__file__)

    steps = """\nadditional notes:
     The report will be saved as an XLSX in the same directory.
     How to run:
      1.a     Login to the VPN
      1.b     Add executable permission using chmod
                  Ex. $chmod 775 {}
      2.a     Run using $./{}
         OR
      2.b     Run using $python3 {}""".format(fName, fName, fName)

    desc = "Scrape Jenkins report and create XLSX report."
    parser = argparse.ArgumentParser(description=desc)
    group = parser.add_mutually_exclusive_group()
    group.add_argument('-s', '--smoke', help="Generate Smoke Report",
                       action="store_true")
    group.add_argument('-r', '--regression', help="Generate Regression Report",
                       action="store_true")
    group.add_argument('-t', '--test', help="Test mode",
                       action="store_true")
    group.add_argument('-b', '--both', help="Generate both smoke and regression reports",
                       action="store_true")
    args = parser.parse_args()

    if len(sys.argv) == 1:
        parser.print_help()
        print(steps)
        sys.exit(0)

    if args.smoke or args.test: 
        automationReport = 'RCO Smoke Tests'
        reportFileName = 'RCO_Smoke_Report'
    elif args.regression:
        automationReport = 'RCO Regression Tests'
        reportFileName = 'RCO_Smokeless_Regression_Report'
    elif args.both:
        processBoth()
    suite = {'nodes': [], 'duration': '', 'total': 0, 'failed': 0,
             'name': automationReport}
    devices = ['Desktop', 'Tablet', 'Mobile']

    nodes = ['Authentication', 'Confirmation', 'Delivery', 'Payment',
             'Review']
    if args.test:
        pprint(args)
        nodes = ['Authentication']
        devices = ['Desktop']

def processBoth():
    procs = []
    p1 = worker('-s')
    procs.append(p1)
    p2 = worker('-r')
    procs.append(p2)
    for p in procs:
        p.start()
    for p in procs:
        p.join()
    sys.exit(0)


def is_venv():
    return (hasattr(sys, 'real_prefix') or
            (hasattr(sys, 'base_prefix') and sys.base_prefix != sys.prefix))


def worker(tag):
    fileName = os.path.basename(__file__)
    return mp.Process(target=subprocess.call, args=([PY, fileName, tag],))

def build_paths():
    global ENV_NAME
    global ENV
    global BIN
    global PIP
    global PY

    ENV = os.path.join('.', ENV_NAME)
    if sys.platform == "win32":
        BIN = os.path.join(ENV, 'Scripts')
    else:
        BIN = os.path.join(ENV, 'bin')
    PIP = os.path.join(BIN,'pip')
    if sys.platform == "win32":
        PY  = os.path.join(BIN,'python')
    else:
        PY  = os.path.join(BIN,'python3')

build_paths()
#janky bootsrap implementation
if not is_venv():
    #create virtual environment
    venv.create(ENV, with_pip=True)
    if sys.platform != "win32":
        subprocess.call([PIP, 'install', '--upgrade','pip'])
        subprocess.call([PIP, 'install', 'selenium'])
        subprocess.call([PIP, 'install', 'openpyxl'])
        subprocess.call([PIP, 'install', 'requests'])
    else:
        subprocess.call([PIP, 'install', '--upgrade','pip'], shell=True)
        subprocess.call([PIP, 'install', 'selenium'], shell=True)
        subprocess.call([PIP, 'install', 'openpyxl'], shell=True)
        subprocess.call([PIP, 'install', 'requests'], shell=True)


def checkCompat():
    compatible = True
    if sys.version_info < (3, 4):
        compatible = False
    elif not hasattr(sys, 'base_prefix'):
        compatible = False
    if not compatible:
        raise ValueError('This script is only for use with '
                         'Python 3.4 or later')

def open_file(filename):
    if sys.platform == "win32":
        os.startfile(filename)
    else:
        opener ="open" if sys.platform == "darwin" else "xdg-open"
        subprocess.call([opener, filename])

def getFeatureFileName(string):
    if string[0] == '(':
        cutoff = 0
        for char in string:
            if char == ')':
                return string[1:cutoff]
            cutoff += 1

def remUnderscore(string):
    final = ''
    for char in string:
        if char != '_':
            final += char
    return final

def cellLineBreak(stringList):
    if len(stringList) == 0:
        return ''
    if len(stringList) == 1:
        return stringList[0]
    final = '"' + stringList[0]
    for i in range(len(stringList)):
        if i == 0:
            continue
        final += '\n' + stringList[i]
    return final + '"'

def duration(string):
    draft = ''
    digits = '1234567890'
    letters = 'hmis'
    for char in string:
        if char in digits:
            draft += char
        if char in letters:
            draft += char
    final = ''
    lastChar = ''
    for char in draft:
        if char == 's':
            if lastChar == 'm':
                final += char
            if lastChar in digits:
                final += char
        elif char in digits:
            if lastChar not in digits:
                final += ' ' + char
            else:
                final += char
        else:
            final += char
        lastChar = char
    return final

def mergeCells(ws, cell, length):
    mergeStart = cell.row
    mergeEnd   = mergeStart + length
    for col in range(1, COLS.index("DURATION")+1):
        ws.merge_cells(start_row = mergeStart,
                start_column     = col,
                end_row          = mergeEnd,
                end_column       = col)

def mergeSheet(ws):
    mergeLen   = 0
    newGroup   = False
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, max_col=1):
        for cell in row:
            addFailedCellFormula(cell)
            if cell.value:
                if not newGroup:
                    start = cell
                    # overwrite failed cell formula
                    addFailedCellFormula(start)
                else:
                    # overwrite failed cell formula
                    addFailedCellFormula(start, mergeLen)
                    mergeCells(ws, start, mergeLen)
                    mergeLen = 0
                    start    = cell
                    newGroup = False
            else:
                if cell.row < ws.max_row:
                    mergeLen += 1
                    newGroup  = True
                else:
                    # overwrite failed cell formula
                    addFailedCellFormula(start, mergeLen)
                    mergeCells(ws, start, mergeLen + 1)

def getAllSuites(ws):
    return [cell for row in ws.iter_rows(min_row=1,
                                         max_row=ws.max_row,
                                         max_col=1)
            for cell in row if str(cell.value).startswith('Suite:') ]

def resizeToFitColumn(ws, columnIndex):
    """ one-based index """
    width = max(list(map(len,
        [str(cell.value) for row in ws.iter_rows(min_row = 1,
                                            max_row = ws.max_row,
                                            min_col = columnIndex,
                                            max_col = columnIndex)
                                            for cell in row])))
    columnName = get_column_letter(columnIndex)
    ws.column_dimensions[columnName].width = width

def resizeRow(ws, rowIndex, multiplier):
    """ one-based index """
    size = units.DEFAULT_ROW_HEIGHT * multiplier
    ws.row_dimensions[rowIndex].height = size

def resizeColumn(ws, columnIndex, width):
    """ one-based index """
    columnName = get_column_letter(columnIndex)
    ws.column_dimensions[columnName].width = width

def initStyles():
    #Red Fill
    global redFill
    redFill = PatternFill(start_color = colors.RED,
                          end_color   = colors.RED,
                          fill_type   = 'solid')

    #Blue Fill
    global blueFill
    blueFill = PatternFill(start_color = colors.BLUE,
                           end_color   = colors.BLUE,
                           fill_type   = 'solid')
    #Green Fill
    global greenFill
    greenFill = PatternFill(start_color = colors.GREEN,
                            end_color = colors.GREEN,
                            fill_type = 'solid')

    #Cyan Fill
    global cyanFill
    cyan = "FF00FFFF"
    cyanFill = PatternFill(start_color = cyan,
                            end_color = cyan,
                            fill_type = 'solid')

    #vertical alignment centered 
    global verCenterAl
    vertCenterAl =  Alignment(vertical="center")

    #horizontal centered alignment
    global horCenterAl
    horCenterAl =  Alignment(horizontal="center")

    #centered everything
    global centerAl
    centerAl =  Alignment(vertical="center", horizontal="center")

    #thin border
    global thinBorder
    bd         = Side(style='thin', color="000000")
    thinBorder = Border(left=bd, top=bd, right=bd, bottom=bd)

    #all cells
    global defaultStyle
    defaultStyle           = NamedStyle(name="default")
    defaultStyle.alignment = vertCenterAl
    defaultStyle.border    = thinBorder

    #blue highlight with white text style
    global blueHighlight
    blueHighlight           = NamedStyle(name="blueHighlight")
    blueHighlight.font      = Font(color=colors.WHITE)
    blueHighlight.alignment = centerAl
    blueHighlight.border    = thinBorder
    blueHighlight.fill      = blueFill

    #cyan highlight with white text style
    global cyanHighlight
    cyanHighlight           = NamedStyle(name="cyanHighlight")
    cyanHighlight.alignment = centerAl
    cyanHighlight.border    = thinBorder
    cyanHighlight.fill      = cyanFill

    #suite header style
    global suiteStyle
    suiteStyle            = NamedStyle(name="Suite")
    suiteStyle.font       = Font(color=colors.WHITE)
    suiteStyle.alignment  = Alignment(vertical="center", horizontal="center")
    suiteStyle.border     = thinBorder
    suiteStyle.fill       = blueFill

    #failed cells style
    global failedStyle
    failedStyle            = NamedStyle(name="Failed")
    failedStyle.alignment = Alignment(vertical="center", wrap_text=True)

    #skipped steps style
    global skippedStyle
    skippedStyle            = NamedStyle(name="Skipped")
    skippedStyle.alignment  = Alignment(vertical="center", horizontal="center", wrap_text=True)

    # data validation for manual testing column
    global manualTestingDv
    manualTestingDv = DataValidation(type="list", formula1='"PASSED,FAILED"', allow_blank=True)
    

def scrapeInfo():
    global suite
    global devices
    global nodes
    for node in nodes:
        for device in devices:
            newNode = {'name': 'Tags=Checkout_' + node + '_Responsive_' +
                       device, 'features': [], 'duration': '', 'total': 0,
                       'failed': 0}
            suite['nodes'].append(newNode)
    driver=wd.Chrome()
    driver.get("http://jenkins.ccp.tm.tmcs/view/RCO/")
    driver.find_element_by_link_text(suite['name']).click()
    for node in suite['nodes']:
        features = []
        driver.find_element_by_link_text(node['name']).click()
        cucumberReport = driver.find_element_by_link_text('Cucumber Reports').click()
        node['duration'] = driver.find_element_by_id('stats-total-duration').text
        featureElements = driver.find_elements_by_partial_link_text('Checkout_')
        featureTexts = []
        for element in featureElements:
            featureTexts.append(element.text)
        for featureText in featureTexts:
            featureElement = driver.find_element_by_link_text(featureText)
            feature = {}
            feature['name'] = featureText
            feature['total'] = int(driver.find_element_by_id('stats-number-scenarios-' +
                                                             featureText).text)
            node['total'] = node['total'] + feature['total']
            suite['total'] = suite['total'] + feature['total']
            feature['duration'] = driver.find_element_by_id('stats-duration-' +
                                                            featureText).text
            featureElement.click()
            #in feature page
            #scraping failures
            failures = driver.find_elements_by_xpath("//div[@class='failed']/span[@class='scenario-keyword']")
            feature['failed'] = len(failures)
            node['failed'] = node['failed'] + feature['failed']
            suite['failed'] = suite['failed'] + feature['failed']
            scenarios = driver.find_elements_by_xpath("//div[@class='failed']/span[@class='scenario-name']")
            steps = driver.find_elements_by_xpath("//div[@class='failed']/span[@class='step-name']")
            feature['failureList'] = []
            feature['failedSteps'] = []
            feature['failures'] = {}
            count = 0
            for num in range(len(failures)):
                if failures[num].text == 'Background:':
                    feature['failureList'].append(failures[num].text)
                    feature['failures'][failures[num].text] = {'scenario': failures[num].text ,
                                                               'step': steps[num].text}
                else:
                    feature['failureList'].append(failures[num].text +
                                                  scenarios[count].text)
                    feature['failures'][failures[num].text +
                                        scenarios[count].text] = {'scenario': ' '.join((failures[num].text + scenarios[count].text).split()),
                                                                  'step': steps[num].text}
                    count += 1
                feature['failedSteps'].append(steps[num].text)

            feature['skipList'] = []
            #scraping skipped
#            skips = driver.find_elements_by_xpath("//div[@class='skipped']/span[@class='step-keyword']")
#            feature['skipped'] = len(skips)
#            node['skipped'] += feature['skipped']
#            suite['skipped'] += feature['skipped']
# WIP
#            scenarios = driver.find_elements_by_xpath("//div[@class='skipped']/span[@class='step-name']")
#            steps = driver.find_elements_by_xpath("//div[@class='skipped']/span[@class='step-name']")
#            feature['skipList'] = []
#            feature['skippedSteps'] = []
#            feature['skips'] = {}
#            count = 0
#            for num in range(len(skips)):
#                if skips[num].text == 'Background:':
#                    feature['skipList'].append(skips[num].text)
#                    feature['skips'][skips[num].text] = {'scenario': skips[num].text ,
#                                                               'step': steps[num].text}
#                else:
#                    feature['skipList'].append(skips[num].text +
#                                                  scenarios[count].text)
#                    feature['skips'][skips[num].text +
#                                        scenarios[count].text] = {'scenario': skips[num].text +
#                                                                  scenarios[count].text,
#                                                                  'step': steps[num].text}
#                    count += 1
#                feature['skippedSteps'].append(steps[num].text)
# WIP end
            #add to scraped info to list
            features.append(feature)
            #leaving feature page
            driver.back()
        node['features'] = features
        driver.back()
        driver.back()
    driver.quit()

def writeToTextFile(suite, reportFileName):
    reportFileName += '.tsv'
    report = open(reportFileName, 'w')
    report.write('Tests' + '\t' + 'Success Rate' + '\t' + '"#\nTests"'
                 + '\t' + '"#\nFailed"' + '\t' + 'Duration' + '\t'
                 + 'Failed Scenarios' + '\t' + 'Failed Steps' + '\n')
    report.write(suite['name'] + '\t' +
                 str(round(100.0*(suite['total']-suite['failed'])/suite['total'],
                           1)) + '%' + '\t' + str(suite['total']) + '\t' +
                 str(suite['failed']) + '\n')
    for node in suite['nodes']:
        if node['total'] == 0:
            percent = 'No Tests'
        else:
            percent = str(round(100.0*(node['total']-node['failed'])/
                                node['total']),1) + '%'
        report.write('"Suite:\n' + node['name'][5:] + '"' + '\t' + percent +
                     '\t' + str(node['total']) + '\t' + str(node['failed']) +
                     '\t' + duration(node['duration']) + '\n')
        for feature in node['features']:
            if feature['total'] == 0:
                percent = 'No Tests'
            else:
                percent = (str(round(100.0*(feature['total'] - feature['failed'])
                               / feature['total'],1)) + '%')
            report.write(getFeatureFileName(feature['name']) + '\t'
                        + percent + '\t' + str(feature['total'])
                        + '\t' + str(feature['failed']) + '\t'
                        + duration(feature['duration']) + '\t')
            count = True
            for failure in feature['failureList']:
                if count:
                    report.write(feature['failures'][failure]['scenario'] +
                                 '\t' + feature['failures'][failure]['step'] +
                                 '\n')
                    count = False
                else:
                    report.write('\t\t\t\t\t' +
                                 feature['failures'][failure]['scenario'] +
                                 '\t' + feature['failures'][failure]['step'] +
                                 '\n')
            if feature['failed'] == 0:
                report.write('\n')
    report.close()
    return

def writeToExcelFile(suite):
    global reportFileName
    reportFileName += '.xlsx'
    wb             = Workbook()
    ws             = wb.active
    header         = ('Tests',
                      'Success Rate',
                      "# Tests",
                      "# Failed",
                      'Duration',
                      'Failed Scenarios',
                      'Failed Steps',
                      'Skipped Steps',
                      'Manual Testing')
    ws.append(header)
    subHeader      = (suite['name'],"")
    ws.append(subHeader)

    for node in suite['nodes']:
        percent=''
        row = ("Suite:  " + node['name'][5:],
               percent,
               '',               
               '',
               duration(node['duration']))

        ws.append(row)

        for feature in node['features']:
            percent = ''
            if feature['total'] == 0:
                pass
            else:
                row = (getFeatureFileName(feature['name']),
                       '',
                       feature['total'],
                       '',
                       duration(feature['duration']), '')
                ws.append(row)

            count = True
            for failure in feature['failureList']:
                if count:
                    ws.cell(row=ws.max_row,
                            column=COLS.index("FAILED_SCENARIOS"),
                            value=feature['failures'][failure]['scenario'][9:])
                    ws.cell(row=ws.max_row,
                            column=COLS.index("FAILED_STEPS"),
                            value=feature['failures'][failure]['step'])
                    count = False
                else:
                    ws.append(
                        {COLS.index("FAILED_SCENARIOS"): feature['failures'][failure]['scenario'][9:],
                        COLS.index("FAILED_STEPS"): feature['failures'][failure]['step']})

            for skip in feature['skipList']:
                if count:
                    ws.cell(row=ws.max_row,
                            column=COLS.index("SKIPPED_SCENARIOS"),
                            value=skip)
                    count = False
                else:
                    ws.append({COLS.index("SKIPPED_SCENARIOS"): skip})
    #merging
    mergeSheet(ws)

    #styling
    #all cells
    for row in ws.rows:
        for cell in row:
            cell.style = defaultStyle

    #first 2 rows
    for row in ws.iter_rows(min_row=1, max_col=ws.max_column, max_row=2):
        for cell in row:
            cell.style = blueHighlight

    #skipped steps rows
    for row in ws.iter_rows(min_row=1, min_col=COLS.index("SKIPPED_SCENARIOS"),
                            max_col=COLS.index("SKIPPED_SCENARIOS"), max_row=2):
        for cell in row:
            cell.style = cyanHighlight

    #test suite rows
    suiteCells = getAllSuites(ws)
    for suiteCell in suiteCells:
        for col in ws.iter_cols(min_row = suiteCell.row,
                                max_col = ws.max_column,
                                max_row = suiteCell.row):
            for cell in col:
                if cell.col_idx == COLS.index("SKIPPED_SCENARIOS"):
                    cell.style = cyanHighlight
                else:
                    cell.style = suiteStyle
                resizeRow(ws, cell.row, 2)

    #style the rate column
    rateColumn = 'B2:B' + str(ws.max_row)
    ws.conditional_formatting.add(rateColumn,
            CellIsRule(operator='==', formula=['1'], fill=greenFill, font=Font(color=colors.BLACK)))
    ws.conditional_formatting.add(rateColumn,
            CellIsRule(operator='!=', formula=['1'], fill=redFill, font=Font(color=colors.WHITE)))

    for row in ws[rateColumn]:
        for cell in row:
#            cell.font = Font(color=colors.WHITE)
            cell.number_format ='0.00%'

    #style up to the duration column
    for row in ws['B:F']:
        for cell in row:
            cell.alignment = centerAl

    #style the skipped scenarios
    loc = get_column_letter(COLS.index("SKIPPED_SCENARIOS"))
    for row in ws[loc + '1:' + loc + str(ws.max_row)]:
        for cell in row:
            cell.alignment = skippedStyle.alignment

    # style the failed scenarios and failed steps columns
    # TODO 'F4:H' to  relative pos
    failedColumns = 'F4:H' + str(ws.max_row)
    for row in ws[failedColumns]:
        for cell in row:
            cell.alignment = failedStyle.alignment

    # style the manual testing column
    manualTestColumn = 'I2:I' + str(ws.max_row)
    for row in ws[manualTestColumn]:
        for cell in row:
            cell.alignment = centerAl
    ws.conditional_formatting.add(manualTestColumn,
            CellIsRule(operator='==', formula=['"PASSED"'], fill=greenFill, font=Font(color=colors.BLACK)))
    ws.conditional_formatting.add(manualTestColumn,
            CellIsRule(operator='==', formula=['"FAILED"'], fill=redFill, font=Font(color=colors.WHITE)))

    #set fixed widths
    ws.column_dimensions['F'].width = 60
    ws.column_dimensions['G'].width = 60
    ws.column_dimensions['H'].width = 60

    #size all the columns
    for  i in range(1, COLS.index("DURATION")):
        resizeToFitColumn(ws, i)

    #resize duration column
    resizeColumn(ws, COLS.index("DURATION"), DURATION_COL_LEN)

    #resize manual testing column
    resizeColumn(ws, COLS.index("MANUAL_TEST"), MAN_TEST_COL_LEN)

    #resize first row
    resizeRow(ws, HEADER_ROW, 2)

    #resize rows with merged cells
    for row in ws.iter_rows(min_row=3, max_col=1, max_row=ws.max_row):
        for cell in row:
            failedScenarioCell = cell.offset(column=COLS.index("FAILED_SCENARIOS")-1)
            failedStepCell = cell.offset(column=COLS.index("FAILED_STEPS")-1)
            skippedScenarioCell = cell.offset(column=COLS.index("SKIPPED_SCENARIOS")-1)
            valueLength = max(map(len,
                                  map(str,(failedScenarioCell.value,
                                           failedStepCell.value,
                                           skippedScenarioCell.value))))
            if(failedScenarioCell.value or skippedScenarioCell.value):
                size   = 1 + (valueLength // LINE_WRAP_LEN)
                resizeRow(ws, cell.row, size)


    # add Success rate formula to suites and features
    for row in ws.iter_rows(min_row=2, min_col=2, max_col=2, max_row=ws.max_row):
        for cell in row:
            tot = cell.offset(column=1)
            ng = cell.offset(column=2)
            cell.value = '=IF({0}=0,0,({0}-{1})/{0})'.format(tot.coordinate, ng.coordinate)

    # add # of tests formula
    # refactor to a function
    for row in ws.iter_rows(min_row=3,
                            min_col=COLS.index("TEST_NO"),
                            max_col=COLS.index("TEST_NO"),
                            max_row=ws.max_row):
        for cell in row:
            leadCell = cell.offset(column=-2)
            if leadCell in suiteCells:
                suiteLength = getSuiteRowLen(leadCell, ws)
                offset = 0 if not suiteLength else 1
                if suiteLength:
                    cell.value = "=SUM({0}{1}:{0}{2})".format(cell.column,
                                                           str(cell.row+offset),
                                                           str(cell.row+suiteLength))
                else:
                    cell.value = 0

    # add # of tests formula
    # refactor to a function
    for row in ws.iter_rows(min_row=3,
                            min_col=COLS.index("FAILED_NO"),
                            max_col=COLS.index("FAILED_NO"),
                            max_row=ws.max_row):
        for cell in row:
            leadCell = cell.offset(column=-3)
            if leadCell in suiteCells:
                suiteLength = getSuiteRowLen(leadCell, ws)
                offset = 0 if not suiteLength else 1
                if suiteLength:
                    cell.value = "=SUM({0}{1}:{0}{2})".format(cell.column,
                                                              str(cell.row+offset),
                                                              str(cell.row+suiteLength))
                else:
                    cell.value = 0

    # add total # tests formula for suite
    # refactor to a function
    testCells = []
    for row in ws.iter_rows(min_row=3,
                            min_col=COLS.index("TEST_NO"),
                            max_col=COLS.index("TEST_NO"),
                            max_row=ws.max_row):
        for cell in row:
            leadCell = cell.offset(column=-2)
            if leadCell in suiteCells:
                testCells.append(cell.coordinate)
    totalTest = ws['C2']
    totalTest.value = '=SUM({})'.format(','.join(testCells))

    # add total # failures formula for suite
    # refactor to a function
    failedCells = []
    for row in ws.iter_rows(min_row=3,
                            min_col=COLS.index("FAILED_NO"),
                            max_col=COLS.index("FAILED_NO"),
                            max_row=ws.max_row):
        for cell in row:
            leadCell = cell.offset(column=-3)
            if leadCell in suiteCells:
                failedCells.append(cell.coordinate)
    totalFailure = ws['D2']
    totalFailure.value = '=SUM({})'.format(','.join(failedCells))

    # data validation for the manual testing column
    ws.add_data_validation(manualTestingDv)
    for row in ws.iter_rows(min_row=4,
                            min_col=COLS.index("MANUAL_TEST"),
                            max_col=COLS.index("MANUAL_TEST"),
                            max_row=ws.max_row):
        for cell in row:
            leadCell = cell.offset(column=-8)
            if not leadCell in suiteCells:
                manualTestingDv.ranges.append(cell.coordinate)

    #save file
    wb.save(reportFileName)

def urlBuilderFromNode(node):
    ''' returns url '''
    global automationReport
    smoke    = 'RCO_Full_Smoke'
    regg     = 'RCO_Full_Regression'
    prefix   = 'http://jenkins.ccp.tm.tmcs/view/RCO/job/'
    type = smoke if 'Smoke' in automationReport else regg
    mid   = '/Browser=chrome,Domain=US,'
    suffix = ',jdk=JDK8u60,restricted_executors=rco/ws/target/cucumber-integration-json-report.json'
    return prefix + type + mid + node + suffix

def getJsonFile(link):
    ''' returns python object '''
    data = {}
    r = requests.get(link)
    if r.status_code != 200:
        print("Error reading JSON on {}, returned {}".format(link, r.status_code))
    try:
        data = r.json()
    except ValueError:
        print("Error reading JSON on {}, Jenkins test might be in progress".format(link))
    return data


def scrapeSkippedFromJSON(data):
    ''' returns dictionary of unique scenarios with skips '''
    out = {}
    for suite in data:
        if 'elements' in suite:
            skipScenarios = []
            for element in suite['elements']:
                for step in element['steps']:
                    if step['result']['status'] == 'skipped':
                        if element['name']:
                            skipScenarios.append(element['name'])
                if skipScenarios:
                    out[suite['name']] = set(skipScenarios)
    return out


def addSkipped(suite):
    for node in suite['nodes']:
        link = urlBuilderFromNode(node['name'])
        table = scrapeSkippedFromJSON(getJsonFile(link))
        for feature in node['features']:
            if feature['name'] in table:
                # step thru every scenario that skipped
                for scenario in table[feature['name']]:
                    sanitizedScenario = ' '.join(scenario.split())
                    # check if skipped scenario is in the fail list
                    if not sanitizedScenario in feature['failureList']:
                        feature['skipList'].append(sanitizedScenario)

def addFailedCellFormula(cell, rowLen=0):
    failedCell = cell.offset(column=3,)
    failedCell.value='=COUNTIF(I{0}:I{1},"Failed" )+COUNTA(F{0}:F{1})'.format(cell.row,
                                                                             cell.row+rowLen)


def getSuiteRowLen(startCell, ws):
    ''' returns row length of the suite '''
    length = 0
    allSuites = getAllSuites(ws)
    for row in ws.iter_rows(min_row=startCell.row+1,
                            min_col=COLS.index("TEST_NAME"),
                            max_col=COLS.index("TEST_NAME"),
                            max_row=ws.max_row):
        for cell in row:
            if cell not in allSuites:
                length += 1
            else:
                return length
    return length            

if __name__ == '__main__':

    # parse args
    arghandler()

    # virtual env installed libraries
    if is_venv():
        from openpyxl import Workbook
        from openpyxl.styles import colors
        from openpyxl.styles import Color, PatternFill, Font, Border, NamedStyle
        from openpyxl.styles import Alignment, Side
        from openpyxl.formatting import Rule
        from openpyxl.formatting.rule import CellIsRule
        from openpyxl.utils import get_column_letter, rows_from_range
        from openpyxl.utils import units
        from openpyxl.worksheet.datavalidation import DataValidation
        from selenium import webdriver as wd
        import json
        import requests
    else:
        print("Not in venv, starting new subprocess call")
        p = worker(sys.argv[1])
        p.start()
        sys.exit(0)


    initStyles()
    checkCompat()
    scrapeInfo()
    addSkipped(suite)
    writeToExcelFile(suite)
    open_file(reportFileName)
